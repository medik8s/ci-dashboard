"""
Flask web server for dashboard
"""

from flask import Flask, render_template, jsonify, request, send_file
from datetime import datetime, timedelta, timezone
from pathlib import Path
import re
import time
import yaml
import threading
import sys
import os
import logging
import io
import json
import urllib.request
import urllib.parse
import urllib.error
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from storage.database import DashboardDatabase
from metrics.calculator import MetricsCalculator
from reports.weekly_report import WeeklyReportGenerator

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Global collection status
collection_status = {
    'running': False,
    'progress': '',
    'error': None,
    'completed_at': None,
    'lock': threading.Lock()
}


def _fbc_short(fbc_image):
    """Extract a short display label from an FBC catalog image reference."""
    if not fbc_image:
        return ''
    if '@' in fbc_image:
        return fbc_image.split('@')[-1][:15]
    if ':' in fbc_image:
        return fbc_image.split(':')[-1][:8]
    return fbc_image


POLARION_BASE = 'https://polarion.engineering.redhat.com/polarion/#/project/OSE/workitem?id='
GITLAB_FBC_PROJECT = 'dragonfly/rhwa-fbc'


def _polarion_url(polarion_id):
    """Build Polarion work item URL from ID."""
    return f"{POLARION_BASE}{polarion_id}" if polarion_id else ''


def _build_fbc_urls(fbc_image, gitlab_fbc_project=GITLAB_FBC_PROJECT):
    """Build Quay, Konflux, and GitLab URLs from an FBC image reference."""
    fbc_tag_url = ''
    fbc_quay_url = ''
    fbc_konflux_url = ''
    fbc_gitlab_url = ''
    snap_name = None
    _iib_empty = {'iib_id': '', 'iib_digest_short': '', 'iib_url': '', 'iib_log_url': '', 'iib_resolved': ''}
    if not fbc_image:
        return {
            'fbc_image_short': '',
            'fbc_image_url': '',
            'fbc_quay_url': '',
            'fbc_konflux_url': '',
            'fbc_gitlab_url': '',
            'snapshot_name': '',
            **_iib_empty,
        }
    if 'quay.io/' in fbc_image:
        repo_path = fbc_image.split('quay.io/')[-1].split('@')[0].split(':')[0]
    elif 'redhat-user-workloads/' in fbc_image:
        repo_path = 'redhat-user-workloads/' + fbc_image.split('redhat-user-workloads/')[-1].split('@')[0].split(':')[0]
    else:
        return {
            'fbc_image_short': _fbc_short(fbc_image),
            'fbc_image_url': '',
            'fbc_quay_url': '',
            'fbc_konflux_url': '',
            'fbc_gitlab_url': '',
            'snapshot_name': '',
            **_iib_empty,
        }
    app_name = ''
    if repo_path:
        fbc_tag_url = f"https://quay.io/repository/{repo_path}?tab=tags"
        if '@' in fbc_image:
            digest = fbc_image.split('@')[-1]
            fbc_quay_url = f"https://quay.io/repository/{repo_path}/manifest/{digest}"
        elif ':' in fbc_image:
            tag = fbc_image.split(':')[-1]
            fbc_quay_url = f"https://quay.io/repository/{repo_path}?tab=tags&tag={tag}"
        else:
            fbc_quay_url = fbc_tag_url
        parts = repo_path.split('/')
        if len(parts) >= 4 and parts[0] == 'redhat-user-workloads':
            tenant = parts[1]
            app_name = parts[-1]
            fbc_konflux_url = f"{KONFLUX_UI}/ns/{tenant}/applications/{app_name}/snapshots"
        if '@' not in fbc_image and ':' in fbc_image:
            commit_sha = fbc_image.split(':')[-1]
            if _FBC_SHA_RE.fullmatch(commit_sha):
                fbc_gitlab_url = f"https://gitlab.cee.redhat.com/{gitlab_fbc_project}/-/commit/{commit_sha}"
                snap_name, snap_app = _resolve_konflux_snapshot(commit_sha, expected_app=app_name or None)
                if snap_name and snap_app:
                    fbc_konflux_url = f"{KONFLUX_UI}/ns/{KONFLUX_NAMESPACE}/applications/{snap_app}/snapshots/{snap_name}"
    snapshot_name = snap_name or ''
    iib_data = _iib_empty
    if app_name and '@' not in fbc_image and ':' in fbc_image:
        sha = fbc_image.split(':')[-1]
        if _FBC_SHA_RE.fullmatch(sha):
            sha = _maybe_expand_sha(sha, app_name)
            iib_data = _resolve_konflux_release(sha, app_name=app_name)
    return {
        'fbc_image_short': _fbc_short(fbc_image),
        'fbc_image_url': fbc_tag_url,
        'fbc_quay_url': fbc_quay_url,
        'fbc_konflux_url': fbc_konflux_url,
        'fbc_gitlab_url': fbc_gitlab_url,
        'snapshot_name': snapshot_name,
        **iib_data,
    }


_FBC_SHA_RE = re.compile(r'[0-9a-fA-F]{7,40}')
_SNAPSHOT_NAME_RE = re.compile(r'rhwa-fbc-\d{3}-\d{8}-\d+(-\d+)?')

KONFLUX_API = "https://api.stone-prod-p02.hjvn.p1.openshiftapps.com:6443"
KONFLUX_NAMESPACE = "rhwa-tenant"
KONFLUX_UI = "https://konflux-ui.apps.stone-prod-p02.hjvn.p1.openshiftapps.com"
QUAY_FBC_REPO_PREFIX = f"redhat-user-workloads/{KONFLUX_NAMESPACE}/rhwa-fbc"
_konflux_token = os.environ.get("KONFLUX_TOKEN", "")
_sha_expansion_cache = {}
_snapshot_cache = {}
_recent_snapshots_cache = {}
_recent_snapshots_ts = 0


def _list_recent_snapshots(app_name, limit=5):
    """List recent Konflux Snapshots for an application."""
    import time as _time
    global _recent_snapshots_ts
    cache_key = app_name
    now = _time.time()
    if cache_key in _recent_snapshots_cache and (now - _recent_snapshots_ts) < 300:
        return _recent_snapshots_cache[cache_key]
    if not _konflux_token or not app_name:
        return []
    selector = f"appstudio.openshift.io/application={app_name}"
    url = (f"{KONFLUX_API}/apis/appstudio.redhat.com/v1alpha1"
           f"/namespaces/{KONFLUX_NAMESPACE}/snapshots"
           f"?labelSelector={urllib.parse.quote(selector, safe='=,')}"
           f"&limit={limit}")
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {_konflux_token}",
        "Accept": "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        results = []
        for item in data.get("items", []):
            labels = item.get("metadata", {}).get("labels", {})
            sha = labels.get("pac.test.appstudio.openshift.io/sha", "")
            name = item.get("metadata", {}).get("name", "")
            created = item.get("metadata", {}).get("creationTimestamp", "")
            if sha and name:
                results.append({
                    'snapshot_name': name,
                    'commit_sha': sha,
                    'created': created,
                    'app_name': app_name,
                })
        _recent_snapshots_cache[cache_key] = results
        _recent_snapshots_ts = now
        return results
    except Exception as exc:
        logger.warning("Konflux snapshot list failed for %s: %s", app_name, exc)
    return []


def _expand_short_sha(short_sha, app_name):
    """Expand a short FBC commit SHA to the full 40-char SHA via Quay tags API."""
    if len(short_sha) < 7:
        return None
    cache_key = (short_sha, app_name)
    if cache_key in _sha_expansion_cache:
        return _sha_expansion_cache[cache_key]
    repo = f"{QUAY_FBC_REPO_PREFIX}/{app_name}"
    url = (f"https://quay.io/api/v1/repository/{repo}/tag/"
           f"?filter_tag_name=like:{short_sha}&limit=5&onlyActiveTags=true")
    try:
        with urllib.request.urlopen(url, timeout=5) as resp:
            for tag in json.loads(resp.read()).get('tags', []):
                name = tag.get('name', '')
                if len(name) == 40 and name.startswith(short_sha):
                    _sha_expansion_cache[cache_key] = name
                    return name
    except (urllib.error.URLError, json.JSONDecodeError, KeyError, ValueError) as exc:
        logger.warning("Quay SHA expansion failed for %s/%s: %s", short_sha, app_name, exc)
    return None


def _maybe_expand_sha(sha, app_name):
    """Return expanded 40-char SHA if input is short, otherwise return as-is."""
    if len(sha) < 40 and app_name:
        return _expand_short_sha(sha, app_name) or sha
    return sha


def _resolve_konflux_snapshot(commit_sha, expected_app=None):
    """Look up a Konflux Snapshot name by FBC commit SHA.

    When expected_app is provided (e.g. 'rhwa-fbc-422'), filters results
    to match the correct OCP version application.
    Returns (snapshot_name, app_name) or (None, None).
    """
    if not _konflux_token or not commit_sha:
        return None, None
    cache_key = (commit_sha, expected_app or '')
    if cache_key in _snapshot_cache:
        return _snapshot_cache[cache_key]
    commit_sha = _maybe_expand_sha(commit_sha, expected_app)
    label = f"pac.test.appstudio.openshift.io/sha={commit_sha}"
    url = (f"{KONFLUX_API}/apis/appstudio.redhat.com/v1alpha1"
           f"/namespaces/{KONFLUX_NAMESPACE}/snapshots"
           f"?labelSelector={urllib.parse.quote(label, safe='=')}")
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {_konflux_token}",
        "Accept": "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        items = data.get("items", [])
        if items and expected_app:
            items = [i for i in items
                     if i.get("metadata", {}).get("labels", {}).get(
                         "appstudio.openshift.io/application", "") == expected_app]
        if items:
            snap = items[0]["metadata"]["name"]
            app = items[0].get("metadata", {}).get("labels", {}).get(
                "appstudio.openshift.io/application", "")
            _snapshot_cache[cache_key] = (snap, app)
            return snap, app
    except Exception as exc:
        logger.warning("Konflux snapshot lookup failed for %s: %s", commit_sha[:8], exc)
    _snapshot_cache[cache_key] = (None, None)
    return None, None


def _resolve_snapshot_name_to_sha(snapshot_name):
    """Resolve a Konflux snapshot name to its commit SHA via the API.

    Queries the Konflux Snapshot CR by metadata.name and extracts the
    pac.test.appstudio.openshift.io/sha label.
    Returns the 40-char hex SHA string, or None if not found.
    """
    if not _konflux_token or not snapshot_name:
        return None
    url = (f"{KONFLUX_API}/apis/appstudio.redhat.com/v1alpha1"
           f"/namespaces/{KONFLUX_NAMESPACE}/snapshots/{snapshot_name}")
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {_konflux_token}",
        "Accept": "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read())
        sha = data.get("metadata", {}).get("labels", {}).get(
            "pac.test.appstudio.openshift.io/sha", "")
        if sha and _FBC_SHA_RE.fullmatch(sha):
            return sha
    except urllib.error.HTTPError as exc:
        if exc.code == 404:
            logger.debug("Snapshot %s not found in Konflux", snapshot_name)
        else:
            logger.warning("Snapshot lookup HTTP %s for %s: %s", exc.code, snapshot_name, exc)
    except Exception as exc:
        logger.warning("Snapshot name lookup failed for %s: %s", snapshot_name, exc)
    return None


_release_cache = {}


def _resolve_konflux_release(commit_sha, app_name=None):
    """Look up IIB build ID and digest from a Konflux Release by FBC commit SHA.

    Returns dict with iib_id, iib_digest_short, iib_url, iib_log_url (all strings).
    """
    cache_key = (commit_sha, app_name or '')
    if cache_key in _release_cache:
        return _release_cache[cache_key]
    empty = {'iib_id': '', 'iib_digest_short': '', 'iib_url': '', 'iib_log_url': '', 'iib_resolved': ''}
    if not _konflux_token or not commit_sha:
        _release_cache[cache_key] = empty
        return empty
    selector = f"pac.test.appstudio.openshift.io/sha={commit_sha}"
    if app_name:
        selector += f",appstudio.openshift.io/application={app_name}"
    url = (f"{KONFLUX_API}/apis/appstudio.redhat.com/v1alpha1"
           f"/namespaces/{KONFLUX_NAMESPACE}/releases"
           f"?labelSelector={urllib.parse.quote(selector, safe='=,')}")
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {_konflux_token}",
        "Accept": "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        items = data.get("items", [])
        if not items:
            _release_cache[cache_key] = empty
            return empty
        artifacts = items[0].get("status", {}).get("artifacts", {})
        idx_map = artifacts.get("index_image", {})
        idx_entry = None
        for ver_key, ver_val in idx_map.items():
            idx_entry = ver_val
            break
        if not idx_entry:
            comps = artifacts.get("components", [])
            if comps:
                idx_entry = comps[0]
        release_name = items[0].get("metadata", {}).get("name", "")
        release_app = items[0].get("metadata", {}).get("labels", {}).get(
            "appstudio.openshift.io/application", app_name or "")
        if idx_entry:
            iib_ref = idx_entry.get("index_image", "")
            iib_resolved = idx_entry.get("index_image_resolved", "")
            iib_id = iib_ref.split(":")[-1] if ":" in iib_ref else ""
            digest = iib_resolved.split("@")[-1] if "@" in iib_resolved else ""
            log_url = ""
            if release_name and release_app:
                log_url = (f"{KONFLUX_UI}/ns/{KONFLUX_NAMESPACE}/applications"
                           f"/{release_app}/releases/{release_name}/artifacts")
            result = {
                'iib_id': iib_id,
                'iib_digest_short': digest[:16] if digest else '',
                'iib_url': iib_ref,
                'iib_log_url': log_url,
                'iib_resolved': iib_resolved,
            }
            _release_cache[cache_key] = result
            return result
    except Exception as exc:
        logger.warning("Konflux release lookup failed for %s: %s", commit_sha[:8], exc)
    _release_cache[cache_key] = empty
    return empty


def _parse_fbc_overrides(data):
    """Parse FBC commit SHA from request data and return (env_overrides, error_msg).

    Accepts hex SHAs (7-40 chars) directly, and resolves Konflux snapshot
    names (e.g. 'rhwa-fbc-422-20260728-080906-000') to their commit SHA
    via the Konflux API.

    Returns ({}, None) when no SHA is provided.
    Returns ({'FBC_COMMIT_SHA': sha}, None) on valid/resolved SHA.
    Returns (None, 'error message') on invalid input or failed resolution.
    """
    if not isinstance(data, dict):
        return {}, None
    fbc_sha = (data.get('fbc_commit_sha') or '').strip()
    if not fbc_sha:
        return {}, None
    if _FBC_SHA_RE.fullmatch(fbc_sha):
        return {'FBC_COMMIT_SHA': fbc_sha}, None
    if _SNAPSHOT_NAME_RE.fullmatch(fbc_sha):
        resolved = _resolve_snapshot_name_to_sha(fbc_sha)
        if resolved:
            return {'FBC_COMMIT_SHA': resolved}, None
        return None, f'Snapshot {fbc_sha!r} not found in Konflux'
    return None, f'Invalid input: expected a hex SHA (7-40 chars) or a snapshot name (rhwa-fbc-NNN-...)'


def _format_export_row(row, empty_placeholder='-'):
    """Shared row formatting for XLSX/CSV/Markdown exports."""
    job_name = row.get('periodic_job') or ''
    build_id = row.get('build_id') or ''
    step_name = row.get('step_name') or ''
    urls = _build_log_urls(job_name, build_id, step_name)
    short_job = job_name.replace('periodic-ci-medik8s-system-tests-main-', '')
    dur_secs = row.get('job_duration')
    if dur_secs and dur_secs > 0:
        h = int(dur_secs) // 3600
        m = (int(dur_secs) % 3600) // 60
        duration_str = f"{h}h {m}m" if h > 0 else f"{m}m"
    else:
        duration_str = empty_placeholder
    run_date_raw = row.get('run_date') or ''
    run_date = run_date_raw.split('T')[0] if run_date_raw else empty_placeholder
    result = row.get('result') or ''
    result_str = 'PASSED' if result == 'passed' else 'FAILED' if result == 'failed' else (result.upper() or '-')
    return {
        'short_job': short_job, 'duration_str': duration_str,
        'run_date': run_date, 'result_str': result_str, **urls,
    }


GCS_BUCKET = 'test-platform-results'
GCS_HOST = 'https://storage.googleapis.com'
GCSWEB_HOST = 'gcsweb-ci.apps.ci.l2s4.p1.openshiftapps.com'


def _build_log_urls(job_name, build_id, step_name, gcs_prefix=None):
    """Build GCS and gcsweb log URLs from job metadata."""
    has_job = bool(job_name and build_id)
    if gcs_prefix:
        gcs_base = f"{GCS_HOST}/{GCS_BUCKET}/{gcs_prefix}"
        gcsweb_base = f"https://{GCSWEB_HOST}/gcs/{GCS_BUCKET}/{gcs_prefix}"
    else:
        gcs_base = f"{GCS_HOST}/{GCS_BUCKET}/logs/{job_name}/{build_id}" if has_job else ''
        gcsweb_base = f"https://{GCSWEB_HOST}/gcs/{GCS_BUCKET}/logs/{job_name}/{build_id}" if has_job else ''
    artifacts_base = f"{gcs_base}/artifacts/{step_name}" if (has_job and step_name) else ''
    return {
        'e2e_log_url': f"{artifacts_base}/e2e-test/build-log.txt" if (has_job and step_name) else '',
        'install_log_url': f"{artifacts_base}/ipi-install-install/build-log.txt" if (has_job and step_name) else '',
        'subscribe_log_url': f"{artifacts_base}/medik8s-operator-subscribe/build-log.txt" if (has_job and step_name) else '',
        'catalog_log_url': f"{artifacts_base}/medik8s-catalogsource/build-log.txt" if (has_job and step_name) else '',
        'artifacts_url': f"{gcsweb_base}/artifacts/{step_name}/gather-must-gather/" if (has_job and step_name) else '',
        'build_log_url': f"{gcs_base}/build-log.txt" if has_job else '',
    }


def run_collection_background(db_path: str, config_file: str = 'config.yaml', days: int = 30):
    """Run data collection in background thread"""
    global collection_status

    try:
        logger.info(f"Starting data collection for {days} days")
        collection_status['progress'] = 'Starting collection...'

        # Load config
        with open(config_file, 'r') as f:
            config = yaml.safe_load(f)

        # Import collector modules
        sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

        # Initialize collector based on type
        collector_type = config['collector']['type']
        logger.info(f"Using collector type: {collector_type}")

        if collector_type == 'reportportal':
            from collectors.reportportal import ReportPortalCollector
            rp_config = config['collector']['reportportal']
            collector = ReportPortalCollector(rp_config)
        elif collector_type == 'prow_mcp':
            from collectors.prow_mcp import ProwMCPCollector
            mcp_config = config['collector']['prow_mcp']
            collector = ProwMCPCollector(mcp_config)
        elif collector_type == 'prow_gcs':
            from collectors.prow_gcs import ProwGCSCollector
            gcs_config = config['collector']['prow_gcs']
            try:
                collector = ProwGCSCollector(gcs_config)
            except Exception as e:
                error_msg = f'Failed to initialize prow_gcs collector: {e}'
                logger.error(error_msg)
                collection_status['error'] = error_msg
                collection_status['running'] = False
                return
        else:
            error_msg = f'Unsupported collector type: {collector_type}'
            logger.error(error_msg)
            collection_status['error'] = error_msg
            collection_status['running'] = False
            return

        # Health check
        logger.info("Running health check...")
        collection_status['progress'] = 'Checking data source...'
        if not collector.health_check():
            error_msg = 'Failed to connect to data source'
            logger.error(error_msg)
            collection_status['error'] = error_msg
            collection_status['running'] = False
            return

        # Calculate date range
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)

        # Get job patterns based on collector type
        versions = config['tracking']['versions']
        platforms = config['tracking']['platforms']

        if collector_type == 'reportportal':
            job_patterns = config['collector']['reportportal']['job_patterns']
            # Expand patterns with version placeholders
            expanded_patterns = []
            for pattern in job_patterns:
                for version in versions:
                    expanded_patterns.append(pattern.replace('{version}', version))
        elif collector_type == 'prow_gcs':
            # prow_gcs uses wildcard patterns, no version expansion needed
            # Support both 'job_patterns' (new) and 'job_names' (legacy)
            prow_gcs_config = config['collector']['prow_gcs']
            expanded_patterns = prow_gcs_config.get('job_patterns') or prow_gcs_config.get('job_names', [])
        elif collector_type == 'prow_mcp':
            # prow_mcp uses exact job names from config
            expanded_patterns = None  # Will use job_names from collector config
        else:
            expanded_patterns = []

        # Collect job runs
        logger.info("Collecting job runs...")
        collection_status['progress'] = 'Collecting job runs...'
        job_runs = collector.collect_job_runs(
            start_date=start_date,
            end_date=end_date,
            job_patterns=expanded_patterns,
            versions=versions,
            platforms=platforms
        )
        logger.info(f"Collected {len(job_runs)} job runs")

        # Collect test results
        collection_status['progress'] = f'Collected {len(job_runs)} job runs, collecting test results...'
        logger.info("Collecting test results (fetching logs for failed tests)...")
        test_results = collector.collect_test_results(
            start_date=start_date,
            end_date=end_date,
            job_patterns=expanded_patterns,
            versions=versions,
            platforms=platforms
        )
        logger.info(f"Collected {len(test_results)} test results")

        # Collect presubmit jobs (if configured)
        presubmit_patterns = config.get('collector', {}).get(
            'prow_gcs', {}
        ).get('presubmit_job_patterns', [])
        if presubmit_patterns and collector_type == 'prow_gcs':
            collection_status['progress'] = 'Collecting presubmit job runs...'
            logger.info("Collecting presubmit job runs...")
            presubmit_job_runs = collector.collect_presubmit_job_runs(
                start_date=start_date,
                end_date=end_date,
                job_patterns=presubmit_patterns,
                versions=versions,
                platforms=platforms,
            )
            logger.info(f"Collected {len(presubmit_job_runs)} presubmit job runs")

            collection_status['progress'] = (
                f'Collected {len(presubmit_job_runs)} presubmit job runs, '
                'collecting presubmit test results...'
            )
            logger.info("Collecting presubmit test results...")
            presubmit_test_results = collector.collect_presubmit_test_results(
                start_date=start_date,
                end_date=end_date,
                job_patterns=presubmit_patterns,
                versions=versions,
                platforms=platforms,
                job_runs=presubmit_job_runs,
            )
            logger.info(f"Collected {len(presubmit_test_results)} presubmit test results")

            job_runs.extend(presubmit_job_runs)
            test_results.extend(presubmit_test_results)

        # Save to database
        collection_status['progress'] = (
            f'Collected {len(job_runs)} job runs and {len(test_results)} test results, '
            'saving to database...'
        )
        logger.info("Saving to database...")
        db = DashboardDatabase(db_path)

        run_id = db.record_collection_start(trigger='web')

        inserted_jobs = db.insert_job_runs(job_runs)
        inserted_tests = db.insert_test_results(test_results)

        # Update job_runs with actual test counts from test_results
        logger.info("Updating job runs with test counts...")
        db.conn.execute("""
            UPDATE job_runs
            SET
                total_tests = (
                    SELECT COUNT(*) FROM test_results
                    WHERE test_results.job_name = job_runs.job_name
                    AND test_results.build_id = job_runs.build_id
                    AND test_results.status != 'skipped'
                ),
                passed_tests = (
                    SELECT COUNT(*) FROM test_results
                    WHERE test_results.job_name = job_runs.job_name
                    AND test_results.build_id = job_runs.build_id
                    AND test_results.status = 'passed'
                ),
                failed_tests = (
                    SELECT COUNT(*) FROM test_results
                    WHERE test_results.job_name = job_runs.job_name
                    AND test_results.build_id = job_runs.build_id
                    AND test_results.status = 'failed'
                ),
                skipped_tests = (
                    SELECT COUNT(*) FROM test_results
                    WHERE test_results.job_name = job_runs.job_name
                    AND test_results.build_id = job_runs.build_id
                    AND test_results.status = 'skipped'
                )
            WHERE EXISTS (
                SELECT 1 FROM test_results
                WHERE test_results.job_name = job_runs.job_name
                AND test_results.build_id = job_runs.build_id
            )
        """)
        db.conn.commit()
        logger.info("Job runs updated with test counts")

        db.record_collection_end(run_id, 'ok', jobs=inserted_jobs, tests=inserted_tests)

        # Close connection after write
        db.close()

        logger.info(f"Collection complete! Inserted {inserted_jobs} job runs and {inserted_tests} test results")
        collection_status['progress'] = f'Complete! Saved {inserted_jobs} job runs and {inserted_tests} test results'
        collection_status['error'] = None
        collection_status['completed_at'] = datetime.now(timezone.utc).isoformat()

    except Exception as e:
        logger.error(f"Collection failed: {e}", exc_info=True)
        try:
            fail_db = DashboardDatabase(db_path)
            fail_db.record_collection_end(run_id, 'failed', error=str(e))
            fail_db.close()
        except Exception:
            pass
        collection_status['error'] = str(e)
        collection_status['progress'] = 'Failed'
        collection_status['completed_at'] = None
    finally:
        logger.info("Collection thread finished")
        collection_status['running'] = False


def create_app(db_path: str, config: dict = None, config_file: str = 'config.yaml'):
    """
    Create Flask application

    Args:
        db_path: Path to SQLite database
        config: Optional Flask configuration
        config_file: Path to YAML configuration file

    Returns:
        Flask app instance
    """
    app = Flask(__name__,
                template_folder=str(Path(__file__).parent / 'templates'),
                static_folder=str(Path(__file__).parent / 'static'))

    # Disable template caching for development
    app.config['TEMPLATES_AUTO_RELOAD'] = True
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
    app.jinja_env.auto_reload = True
    app.jinja_env.cache = {}

    if config:
        app.config.update(config)

    # Load tracking config for blocklist and job schedules
    blocklist = []
    job_schedules = {}
    try:
        with open(config_file, 'r') as f:
            yaml_config = yaml.safe_load(f)
            blocklist = yaml_config.get('tracking', {}).get('blocklist', [])
            job_schedules = yaml_config.get('collector', {}).get('prow_gcs', {}).get('job_schedules', {})
    except Exception as e:
        print(f"Warning: Could not load config: {e}")

    # Initialize database and calculator
    db = DashboardDatabase(db_path)
    calculator = MetricsCalculator(db, blocklist=blocklist)
    report_generator = WeeklyReportGenerator(db, blocklist=blocklist)

    global collection_status
    try:
        latest = db.execute_query(
            "SELECT MAX(timestamp) as latest FROM job_runs"
        )
        if latest and latest[0].get('latest'):
            collection_status['completed_at'] = latest[0]['latest']
    except Exception as e:
        logger.debug("Could not read last sync time from DB: %s", e)

    # Check if AI analysis is enabled (default: False for production safety)
    enable_ai = os.environ.get('ENABLE_AI_ANALYSIS', 'false').lower() == 'true'

    def get_latest_version():
        """
        Get the latest version from database.
        Returns the highest version number (e.g., "4.22" if both "4.21" and "4.22" exist)
        """
        query = "SELECT DISTINCT version FROM job_runs ORDER BY version DESC LIMIT 1"
        result = db.execute_query(query)
        return result[0]['version'] if result else None

    def normalize_version(version):
        """
        Normalize version parameter: if empty/None, return latest version.
        This prevents statistically invalid aggregation across different versions.
        """
        if not version or version == '':
            return get_latest_version()
        return version

    @app.route('/healthz')
    def healthz():
        return 'ok', 200

    @app.route('/')
    def index():
        """Render main dashboard page"""
        # Check if database needs data collection
        global collection_status

        # Check if database is empty or has no recent data
        try:
            # Query for recent data (last 7 days)
            recent_count = db.execute_query(
                "SELECT COUNT(*) as cnt FROM job_runs WHERE timestamp >= datetime('now', '-7 days')"
            )
            needs_collection = recent_count[0]['cnt'] == 0 if recent_count else True

            # Auto-trigger collection if needed and not already running
            if needs_collection and not collection_status['running']:
                with collection_status['lock']:
                    if not collection_status['running']:
                        collection_status['running'] = True
                        collection_status['progress'] = 'Initializing...'
                        collection_status['error'] = None

                        # Start background thread
                        thread = threading.Thread(
                            target=run_collection_background,
                            args=(db_path, config_file, 30),
                            daemon=True
                        )
                        thread.start()

        except Exception as e:
            print(f"Error checking database status: {e}")

        return render_template('dashboard.html', enable_ai=enable_ai,
                               gitlab_fbc_project=GITLAB_FBC_PROJECT)

    @app.route('/logs')
    def view_logs():
        """Display test logs in a new page"""
        log_content = request.args.get('content', '')
        test_name = request.args.get('test', 'Test Log')

        html = f'''
        <!DOCTYPE html>
        <html>
        <head>
            <title>{test_name} - Logs</title>
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    margin: 0;
                    padding: 20px;
                    background: #f8fafc;
                }}
                .container {{
                    max-width: 1200px;
                    margin: 0 auto;
                    background: white;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    overflow: hidden;
                }}
                .header {{
                    background: #1e40af;
                    color: white;
                    padding: 20px;
                    font-size: 18px;
                    font-weight: 600;
                }}
                .content {{
                    padding: 20px;
                }}
                pre {{
                    background: #1e293b;
                    color: #e2e8f0;
                    padding: 20px;
                    border-radius: 6px;
                    overflow-x: auto;
                    font-family: 'Consolas', 'Monaco', 'Courier New', monospace;
                    font-size: 13px;
                    line-height: 1.6;
                    white-space: pre-wrap;
                    word-wrap: break-word;
                }}
                .error {{
                    color: #fca5a5;
                }}
                .info {{
                    color: #93c5fd;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">{test_name}</div>
                <div class="content">
                    <pre>{log_content}</pre>
                </div>
            </div>
        </body>
        </html>
        '''
        return html

    @app.route('/api/collection-status')
    def api_collection_status():
        """Get current collection status"""
        global collection_status
        return jsonify({
            'running': collection_status['running'],
            'progress': collection_status['progress'],
            'error': collection_status['error'],
            'completed_at': collection_status['completed_at']
        })

    @app.route('/api/trigger-collection', methods=['POST'])
    def api_trigger_collection():
        """Manually trigger data collection"""
        global collection_status

        json_data = request.get_json(silent=True) or {}
        days = json_data.get('days', 30)

        with collection_status['lock']:
            if collection_status['running']:
                return jsonify({'error': 'Collection already running'}), 409

            collection_status['running'] = True
            collection_status['progress'] = 'Initializing...'
            collection_status['error'] = None
            collection_status['completed_at'] = None

            # Start background thread
            thread = threading.Thread(
                target=run_collection_background,
                args=(db_path, config_file, days),
                daemon=True
            )
            thread.start()

        return jsonify({'status': 'started'})

    @app.route('/api/metadata')
    def api_metadata():
        """Get available versions and platforms from database"""
        query_versions = "SELECT DISTINCT version FROM job_runs ORDER BY version DESC"
        query_platforms = "SELECT DISTINCT platform FROM job_runs ORDER BY platform"

        versions = [row['version'] for row in db.execute_query(query_versions)]
        platforms = [row['platform'] for row in db.execute_query(query_platforms)]

        return jsonify({
            'versions': versions,
            'platforms': platforms
        })

    @app.route('/api/summary')
    def api_summary():
        """Get summary statistics"""
        days = request.args.get('days', 7, type=int)
        version = normalize_version(request.args.get('version'))
        platform = request.args.get('platform')
        stats = calculator.get_summary_stats(days=days, version=version, platform=platform)
        return jsonify(stats)

    @app.route('/api/collector-status')
    def api_collector_status():
        status = db.get_collection_status()
        if not status:
            return jsonify({
                'status': 'unknown',
                'message': 'No collection has run yet',
                'freshness': 'stale',
                'last_finished': None,
                'jobs_collected': 0,
                'tests_collected': 0,
            })

        finished = status.get('finished_at')
        current_status = status.get('status', 'unknown')
        hours_ago = None
        freshness = 'stale'

        if finished:
            try:
                fin_dt = datetime.fromisoformat(finished)
                hours_ago = (datetime.utcnow() - fin_dt).total_seconds() / 3600
                if hours_ago < 26:
                    freshness = 'fresh'
                elif hours_ago < 50:
                    freshness = 'warning'
                else:
                    freshness = 'stale'
            except (ValueError, TypeError):
                pass
        elif current_status == 'running':
            freshness = 'running'

        return jsonify({
            'status': current_status,
            'freshness': freshness,
            'last_started': status.get('started_at'),
            'last_finished': finished,
            'hours_ago': round(hours_ago, 1) if hours_ago is not None else None,
            'jobs_collected': status.get('jobs_collected', 0),
            'tests_collected': status.get('tests_collected', 0),
            'error_message': status.get('error_message'),
            'trigger': status.get('trigger', 'unknown'),
        })

    @app.route('/api/test-results')
    def api_test_results():
        """Get enriched test results with all 17 spreadsheet columns"""
        days = request.args.get('days', 30, type=int)
        operator = request.args.get('operator')
        version = normalize_version(request.args.get('version'))

        rows = db.get_enriched_test_results(days=days, operator=operator, version=version)

        results = []
        for row in rows:
            step_name = row.get('step_name') or ''
            job_name = row.get('periodic_job') or ''
            build_id = row.get('build_id') or ''
            urls = _build_log_urls(job_name, build_id, step_name)

            fbc_image = row.get('fbc_image') or ''
            fbc_urls = _build_fbc_urls(fbc_image)

            polarion_id = row.get('polarion_id') or ''
            polarion_url = _polarion_url(polarion_id)

            results.append({
                'test_name': row.get('test_name'),
                'test_description': row.get('test_description'),
                'operator': row.get('operator'),
                'result': row.get('result'),
                'periodic_job': job_name,
                'build_id': build_id,
                'run_date': row.get('run_date'),
                'job_duration': row.get('job_duration'),
                'version': row.get('version'),
                'platform': row.get('platform'),
                'ocp_version': row.get('ocp_version'),
                'csv_version': row.get('csv_version'),
                'fbc_image': fbc_image,
                **fbc_urls,
                'prow_url': row.get('job_url') or '',
                **urls,
                'polarion_id': polarion_id,
                'polarion_url': polarion_url,
                'classification': row.get('manual_classification'),
                'jira_key': row.get('jira_issue_key'),
            })

        return jsonify({'results': results, 'total': len(results)})

    @app.route('/api/operator-stats')
    def api_operator_stats():
        """Get per-operator pass/fail counts"""
        days = request.args.get('days', 30, type=int)
        version = normalize_version(request.args.get('version'))
        stats = db.get_operator_stats(days=days, version=version)
        return jsonify({'operators': stats})

    @app.route('/api/regressions')
    def api_regressions():
        """Detect regressions by comparing latest vs previous periodic run per job."""
        operator = request.args.get('operator')
        version = normalize_version(request.args.get('version'))
        rows = db.get_regressions(operator=operator, version=version)

        regressions = []
        fixes = []
        persistent = []
        new_failures = []

        for row in rows:
            change_type = row.get('change_type', 'stable')
            if change_type == 'stable':
                continue

            step_name = row.get('step_name') or ''
            job_name = row.get('job_name') or ''
            build_id = row.get('build_id') or ''
            urls = _build_log_urls(job_name, build_id, step_name)

            curr_fbc = row.get('fbc_image') or ''
            prev_fbc = row.get('prev_fbc_image') or ''
            curr_fbc_urls = _build_fbc_urls(curr_fbc)
            prev_fbc_urls = _build_fbc_urls(prev_fbc)

            entry = {
                'test_name': row.get('test_name'),
                'operator': row.get('operator'),
                'prev_status': row.get('prev_status'),
                'curr_status': row.get('curr_status'),
                'error_message': row.get('error_message'),
                'job_name': job_name,
                'build_id': build_id,
                'curr_run_date': row.get('curr_run_date'),
                'prev_run_date': row.get('prev_run_date'),
                'curr_job_url': row.get('curr_job_url') or '',
                'prev_job_url': row.get('prev_job_url') or '',
                'ocp_version': row.get('ocp_version'),
                'csv_version': row.get('csv_version'),
                'prev_ocp_version': row.get('prev_ocp_version'),
                'prev_csv_version': row.get('prev_csv_version'),
                'fbc_image_short': curr_fbc_urls.get('fbc_image_short', ''),
                'fbc_konflux_url': curr_fbc_urls.get('fbc_konflux_url', ''),
                'fbc_snapshot_name': curr_fbc_urls.get('snapshot_name', ''),
                'prev_fbc_image_short': prev_fbc_urls.get('fbc_image_short', ''),
                'prev_fbc_konflux_url': prev_fbc_urls.get('fbc_konflux_url', ''),
                'prev_fbc_snapshot_name': prev_fbc_urls.get('snapshot_name', ''),
                'prev_error_message': row.get('prev_error_message'),
                'polarion_id': row.get('polarion_id'),
                'polarion_url': _polarion_url(row.get('polarion_id') or ''),
                **urls,
            }

            if change_type == 'regression':
                regressions.append(entry)
            elif change_type == 'fix':
                fixes.append(entry)
            elif change_type == 'persistent':
                persistent.append(entry)
            elif change_type == 'new_failure':
                new_failures.append(entry)

        return jsonify({
            'regressions': regressions,
            'fixes': fixes,
            'persistent': persistent,
            'new_failures': new_failures,
            'counts': {
                'regressions': len(regressions),
                'fixes': len(fixes),
                'persistent': len(persistent),
                'new_failures': len(new_failures),
            }
        })

    @app.route('/api/job-runs')
    def api_job_runs():
        """Get job run history with enriched metadata"""
        days = request.args.get('days', 30, type=int)
        operator = request.args.get('operator')
        version = normalize_version(request.args.get('version'))
        rows = db.get_job_run_history(days=days, operator=operator, version=version)

        runs = []
        for row in rows:
            step_name = row.get('step_name') or ''
            job_name = row.get('job_name') or ''
            build_id = row.get('build_id') or ''
            urls = _build_log_urls(job_name, build_id, step_name)
            fbc_image = row.get('fbc_image') or ''

            runs.append({
                'job_name': job_name,
                'build_id': build_id,
                'status': row.get('status'),
                'run_date': row.get('run_date'),
                'duration': row.get('duration_seconds'),
                'version': row.get('version'),
                'platform': row.get('platform'),
                'ocp_version': row.get('ocp_version'),
                'csv_version': row.get('csv_version'),
                'fbc_image': fbc_image,
                **_build_fbc_urls(fbc_image),
                'step_name': step_name,
                'total_tests': row.get('total_tests'),
                'passed_tests': row.get('passed_tests'),
                'failed_tests': row.get('failed_tests'),
                'pass_rate': row.get('pass_rate'),
                'prow_url': row.get('job_url') or '',
                'failure_reason': row.get('failure_reason') or '',
                'failed_step': row.get('failed_step') or '',
                'failure_category': row.get('failure_category') or '',
                **urls,
            })

        return jsonify({'job_runs': runs})

    @app.route('/api/fbc-summary')
    def api_fbc_summary():
        """Get FBC validation summary: per-FBC pass/fail across operators"""
        days = request.args.get('days', 30, type=int)
        version = normalize_version(request.args.get('version'))
        rows = db.get_job_run_history(days=days, version=version)

        fbc_map = {}
        for row in rows:
            fbc_image = row.get('fbc_image') or ''
            if not fbc_image:
                continue
            tag = fbc_image.split(':')[-1] if ':' in fbc_image else fbc_image
            if not re.fullmatch(r'[0-9a-fA-F]{7,40}', tag):
                continue
            short = tag[:7]
            if short not in fbc_map:
                fbc_urls = _build_fbc_urls(fbc_image)
                fbc_map[short] = {
                    'fbc_short': short,
                    'fbc_full': tag,
                    'fbc_image': fbc_image,
                    **fbc_urls,
                    'operators': {},
                    'latest_date': None,
                }
            entry = fbc_map[short]
            step_name = row.get('step_name') or ''
            job_name = row.get('job_name') or ''
            build_id = row.get('build_id') or ''
            op = ''
            m = re.search(r'e2e-([a-z]+)', step_name or job_name)
            if m:
                op = m.group(1).upper()
            if not op:
                continue
            urls = _build_log_urls(job_name, build_id, step_name)
            run_date = row.get('run_date') or ''
            if run_date and (not entry['latest_date'] or run_date > entry['latest_date']):
                entry['latest_date'] = run_date
            step_short = re.sub(r'^e2e-', '', step_name) if step_name else op.lower()
            op_key = step_short
            if op_key in entry['operators']:
                existing = entry['operators'][op_key]
                if run_date > (existing.get('run_date') or ''):
                    pass
                else:
                    continue
            dur = row.get('duration_seconds')
            dur_str = ''
            if dur and dur > 0:
                h = int(dur) // 3600
                mins = (int(dur) % 3600) // 60
                dur_str = f"{h}h {mins}m" if h > 0 else f"{mins}m"
            entry['operators'][op_key] = {
                'operator': op,
                'status': row.get('status') or 'unknown',
                'passed_tests': row.get('passed_tests') or 0,
                'failed_tests': row.get('failed_tests') or 0,
                'total_tests': row.get('total_tests') or 0,
                'duration': dur_str,
                'run_date': run_date,
                'job_name': job_name,
                'build_id': build_id,
                'prow_url': f"https://prow.ci.openshift.org/view/gs/test-platform-results/logs/{job_name}/{build_id}" if job_name and build_id else '',
                'e2e_log_url': urls.get('e2e_log_url', ''),
                'install_log_url': urls.get('install_log_url', ''),
                'subscribe_log_url': urls.get('subscribe_log_url', ''),
                'catalog_log_url': urls.get('catalog_log_url', ''),
                'artifacts_url': urls.get('artifacts_url', ''),
                'build_log_url': urls.get('build_log_url', ''),
                'ocp_version': row.get('ocp_version') or '',
                'version': row.get('version') or '',
                'platform': row.get('platform') or '',
                'csv_version': row.get('csv_version') or '',
                'fbc_image': fbc_image,
                'fbc_image_short': entry.get('fbc_image_short', ''),
                'fbc_quay_url': entry.get('fbc_quay_url', ''),
                'fbc_konflux_url': entry.get('fbc_konflux_url', ''),
                'fbc_gitlab_url': entry.get('fbc_gitlab_url', ''),
                'iib_id': entry.get('iib_id', ''),
                'iib_digest_short': entry.get('iib_digest_short', ''),
                'iib_url': entry.get('iib_url', ''),
                'iib_log_url': entry.get('iib_log_url', ''),
                'iib_resolved': entry.get('iib_resolved', ''),
                'snapshot_name': entry.get('snapshot_name', ''),
                'failure_category': row.get('failure_category') or '',
            }

        all_jobs = {}
        for e in fbc_map.values():
            for key, val in e['operators'].items():
                if key not in all_jobs and val.get('job_name'):
                    all_jobs[key] = val['job_name']
        all_ops = sorted(all_jobs.keys())
        summaries = []
        for short in sorted(fbc_map.keys(), key=lambda k: fbc_map[k].get('latest_date') or '', reverse=True):
            e = fbc_map[short]
            passed = sum(1 for o in e['operators'].values() if o['status'] == 'passed')
            failed = sum(1 for o in e['operators'].values() if o['status'] == 'failed')
            total = len(e['operators'])
            summaries.append({
                'fbc_short': e['fbc_short'],
                'fbc_full': e['fbc_full'],
                'fbc_quay_url': e.get('fbc_quay_url', ''),
                'fbc_konflux_url': e.get('fbc_konflux_url', ''),
                'fbc_gitlab_url': e.get('fbc_gitlab_url', ''),
                'iib_id': e.get('iib_id', ''),
                'iib_digest_short': e.get('iib_digest_short', ''),
                'iib_url': e.get('iib_url', ''),
                'iib_log_url': e.get('iib_log_url', ''),
                'iib_resolved': e.get('iib_resolved', ''),
                'snapshot_name': e.get('snapshot_name', ''),
                'latest_date': (e['latest_date'] or '').split('T')[0],
                'passed': passed,
                'failed': failed,
                'total': total,
                'not_run': len(all_ops) - total,
                'operators': e['operators'],
            })

        existing_shas = set()
        for s in summaries:
            if s.get('fbc_full'):
                existing_shas.add(s['fbc_full'][:7])
        seen_apps = set()
        for e in fbc_map.values():
            repo = (e.get('fbc_quay_url') or '').split('/')
            for part in repo:
                if part.startswith('rhwa-fbc-'):
                    seen_apps.add(part)
        for app_name in (seen_apps or {'rhwa-fbc-422'}):
            recent = _list_recent_snapshots(app_name, limit=5)
            for snap in recent:
                short = snap['commit_sha'][:7]
                if short not in existing_shas:
                    existing_shas.add(short)
                    fbc_urls = _build_fbc_urls(
                        f"quay.io/{QUAY_FBC_REPO_PREFIX}/{app_name}:{snap['commit_sha']}")
                    summaries.append({
                        'fbc_short': short,
                        'fbc_full': snap['commit_sha'],
                        'fbc_quay_url': fbc_urls.get('fbc_quay_url', ''),
                        'fbc_konflux_url': fbc_urls.get('fbc_konflux_url', ''),
                        'fbc_gitlab_url': fbc_urls.get('fbc_gitlab_url', ''),
                        'iib_id': fbc_urls.get('iib_id', ''),
                        'iib_digest_short': fbc_urls.get('iib_digest_short', ''),
                        'iib_url': fbc_urls.get('iib_url', ''),
                        'iib_log_url': fbc_urls.get('iib_log_url', ''),
                        'iib_resolved': fbc_urls.get('iib_resolved', ''),
                        'snapshot_name': fbc_urls.get('snapshot_name', '') or snap['snapshot_name'],
                        'latest_date': snap['created'].split('T')[0],
                        'passed': 0, 'failed': 0, 'total': 0,
                        'not_run': len(all_ops),
                        'operators': {},
                        'has_runs': False,
                    })

        sched_data = {}
        for key, sched in job_schedules.items():
            sched_data[key] = {
                'label': sched.get('label', key),
                'variant': sched.get('variant', ''),
                'day': sched.get('day', ''),
            }

        return jsonify({
            'fbc_summaries': summaries,
            'all_operators': all_ops,
            'all_jobs': all_jobs,
            'job_schedules': sched_data,
        })

    @app.route('/api/job-schedules')
    def api_job_schedules():
        sched_data = {}
        for key, sched in job_schedules.items():
            sched_data[key] = {
                'label': sched.get('label', key),
                'variant': sched.get('variant', ''),
                'day': sched.get('day', ''),
            }
        return jsonify({'job_schedules': sched_data})

    @app.route('/api/presubmit-results')
    def api_presubmit_results():
        """Get presubmit test results"""
        days = request.args.get('days', 30, type=int)
        operator = request.args.get('operator')
        version = normalize_version(request.args.get('version'))
        rows = db.get_presubmit_test_results(days=days, operator=operator, version=version)

        results = []
        for row in rows:
            step_name = row.get('step_name') or ''
            job_name = row.get('job_name') or ''
            build_id = row.get('build_id') or ''
            urls = _build_log_urls(job_name, build_id, step_name, gcs_prefix=row.get('gcs_prefix'))
            pr_number = row.get('pr_number') or row.get('jr_pr_number')

            fbc_image = row.get('fbc_image') or ''
            fbc_urls = _build_fbc_urls(fbc_image)

            polarion_id = row.get('polarion_id') or ''
            polarion_url = _polarion_url(polarion_id)

            results.append({
                'test_name': row.get('test_name'),
                'test_description': row.get('test_description'),
                'operator': row.get('operator'),
                'result': row.get('result'),
                'polarion_id': polarion_id,
                'polarion_url': polarion_url,
                'pr_number': pr_number,
                'pr_author': row.get('pr_author'),
                'pr_repo': row.get('pr_repo'),
                'job_name': job_name,
                'build_id': build_id,
                'run_date': row.get('run_date'),
                'duration': row.get('job_duration'),
                'version': row.get('version'),
                'platform': row.get('platform'),
                'ocp_version': row.get('ocp_version'),
                'csv_version': row.get('csv_version'),
                'fbc_image': fbc_image,
                **fbc_urls,
                'step_name': step_name,
                'prow_url': row.get('job_url') or '',
                **urls,
            })

        return jsonify({'results': results})

    @app.route('/api/presubmit-job-runs')
    def api_presubmit_job_runs():
        """Get presubmit job run history"""
        days = request.args.get('days', 30, type=int)
        operator = request.args.get('operator')
        version = normalize_version(request.args.get('version'))
        rows = db.get_presubmit_job_runs(days=days, operator=operator, version=version)

        runs = []
        for row in rows:
            step_name = row.get('step_name') or ''
            job_name = row.get('job_name') or ''
            build_id = row.get('build_id') or ''
            urls = _build_log_urls(job_name, build_id, step_name, gcs_prefix=row.get('gcs_prefix'))

            runs.append({
                'job_name': job_name,
                'build_id': build_id,
                'status': row.get('status'),
                'run_date': row.get('run_date'),
                'duration': row.get('duration_seconds'),
                'version': row.get('version'),
                'platform': row.get('platform'),
                'ocp_version': row.get('ocp_version'),
                'step_name': step_name,
                'total_tests': row.get('total_tests'),
                'passed_tests': row.get('passed_tests'),
                'failed_tests': row.get('failed_tests'),
                'pass_rate': row.get('pass_rate'),
                'pr_number': row.get('pr_number'),
                'pr_author': row.get('pr_author'),
                'pr_repo': row.get('pr_repo'),
                'prow_url': row.get('job_url') or '',
                **urls,
            })

        return jsonify({'job_runs': runs})

    @app.route('/api/trend')
    def api_trend():
        """Get overall pass rate trend"""
        days = request.args.get('days', 30, type=int)
        version = normalize_version(request.args.get('version'))
        platform = request.args.get('platform')

        trend = calculator.get_overall_trend(
            days=days,
            version=version,
            platform=platform
        )
        return jsonify(trend)

    @app.route('/api/test-rankings')
    def api_test_rankings():
        """Get test rankings (worst performers)"""
        days = request.args.get('days', 30, type=int)
        version = normalize_version(request.args.get('version'))
        platform = request.args.get('platform')
        limit = request.args.get('limit', 20, type=int)

        rankings = calculator.get_test_rankings(
            days=days,
            version=version,
            platform=platform,
            limit=limit
        )
        return jsonify(rankings)

    @app.route('/api/version-comparison')
    def api_version_comparison():
        """Compare pass rates across versions"""
        days = request.args.get('days', 30, type=int)
        comparison = calculator.get_version_comparison(days=days)
        return jsonify(comparison)

    @app.route('/api/platform-comparison')
    def api_platform_comparison():
        """Compare pass rates across platforms"""
        days = request.args.get('days', 30, type=int)
        version = normalize_version(request.args.get('version'))

        comparison = calculator.get_platform_comparison(
            days=days,
            version=version
        )
        return jsonify(comparison)

    @app.route('/api/weekly-report')
    def api_weekly_report():
        """Get weekly platform breakdown report"""
        current_days = request.args.get('current_days', 7, type=int)
        previous_days = request.args.get('previous_days', 7, type=int)
        version = normalize_version(request.args.get('version'))
        top = request.args.get('top', 10, type=int)

        # Get platform comparison
        comparison = report_generator.get_platform_week_over_week(
            current_week_days=current_days,
            previous_week_days=previous_days,
            version=version
        )

        # Get top failing tests
        top_tests = calculator.get_test_rankings(days=current_days, version=version, limit=top)

        # Get overall summary
        summary = calculator.get_summary_stats(days=current_days, version=version)

        return jsonify({
            'comparison': comparison,
            'top_tests': top_tests,
            'summary': summary
        })

    @app.route('/api/platform-tests')
    def api_platform_tests():
        """Get test results for a specific platform"""
        platform = request.args.get('platform')
        days = request.args.get('days', 7, type=int)
        version = normalize_version(request.args.get('version'))

        if not platform:
            return jsonify({'error': 'Platform parameter is required'}), 400

        # Get test rankings for this platform
        tests = calculator.get_test_rankings(days=days, version=version, platform=platform, limit=100)

        # Get platform-specific summary
        summary = calculator.get_summary_stats(days=days, platform=platform, version=version)

        return jsonify({
            'platform': platform,
            'tests': tests,
            'summary': summary,
            'days': days
        })

    @app.route('/api/test-error-by-platform')
    def api_test_error_by_platform():
        """Get latest error for a specific test on a specific platform"""
        test_name = request.args.get('test_name')
        version = normalize_version(request.args.get('version'))
        platform = request.args.get('platform')
        days = request.args.get('days', 30, type=int)

        if not test_name or not platform:
            return jsonify({'error': 'test_name and platform parameters are required'}), 400

        # Query for most recent failure on this platform
        query = """
            SELECT
                error_message,
                timestamp,
                job_name,
                build_id,
                job_url,
                platform
            FROM test_results
            WHERE test_name = ?
            AND platform = ?
            AND status = 'failed'
            AND error_message IS NOT NULL
            AND timestamp >= datetime('now', ? || ' days')
        """

        params = [test_name, platform, f'-{days}']

        if version:
            query += " AND version = ?"
            params.append(version)

        query += " ORDER BY timestamp DESC LIMIT 1"

        result = db.execute_query(query, params)

        if result:
            return jsonify(result[0])
        else:
            return jsonify({'error': 'No error found for this test/platform combination'}), 404

    @app.route('/api/get-affected-platforms', methods=['POST'])
    def api_get_affected_platforms():
        """Get all platforms affected by a test failure"""
        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        test_name = data.get('test_name')
        version = data.get('version')
        days = data.get('days', 7)

        if not all([test_name, version]):
            return jsonify({'error': 'Missing required fields: test_name, version'}), 400

        platforms = db.get_affected_platforms(test_name, version, days)
        return jsonify({'platforms': platforms})

    @app.route('/api/jira/create', methods=['POST'])
    def api_create_jira():
        """Create or find existing Jira issue for a test failure"""
        from integrations import get_jira_integration

        jira = get_jira_integration()
        if not jira:
            return jsonify({
                'status': 'disabled',
                'message': 'Jira integration not configured. Set JIRA_API_TOKEN environment variable.'
            })

        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        # Required fields
        test_name = data.get('test_name')
        version = data.get('version')
        platforms = data.get('platforms', [])

        if not all([test_name, version]):
            return jsonify({'error': 'Missing required fields: test_name, version'}), 400

        # If no platforms provided, use single platform from old API
        if not platforms:
            platform = data.get('platform')
            if platform:
                platforms = [platform]

        # Optional fields
        test_description = data.get('test_description', '')
        error_message = data.get('error_message', '')
        job_url = data.get('job_url', '')
        failure_rate = data.get('failure_rate', 0.0)
        runs = data.get('runs', 0)
        failures = data.get('failures', 0)

        # Check for existing issue first (search by test_name + version only)
        existing_issue = jira.search_existing_issue(test_name, version)
        if existing_issue:
            issue_key = existing_issue.get('key')
            issue_url = jira.get_issue_url(issue_key)
            # Save to database (applies to all platforms)
            db.save_jira_issue(test_name, version, jira_issue_key=issue_key)
            return jsonify({
                'status': 'existing',
                'issue_key': issue_key,
                'issue_url': issue_url,
                'message': f'Found existing issue: {issue_key}'
            })

        # Create new issue
        issue_key = jira.create_issue(
            test_name=test_name,
            test_description=test_description,
            version=version,
            platforms=platforms,
            error_message=error_message,
            job_url=job_url,
            failure_rate=failure_rate,
            runs=runs,
            failures=failures
        )

        if issue_key:
            issue_url = jira.get_issue_url(issue_key)
            # Save to database (applies to all platforms)
            db.save_jira_issue(test_name, version, jira_issue_key=issue_key)
            return jsonify({
                'status': 'created',
                'issue_key': issue_key,
                'issue_url': issue_url,
                'message': f'Created new issue: {issue_key}'
            })
        else:
            return jsonify({'error': 'Failed to create Jira issue'}), 500

    TRIGGER_COOLDOWN_SECONDS = 300

    @app.route('/api/trigger-job/map')
    def api_trigger_job_map():
        from integrations.gangway_client import get_operator_job_map, get_all_triggerable_jobs
        from integrations import get_gangway_client
        gangway = get_gangway_client()
        job_map = get_operator_job_map()
        return jsonify({
            'enabled': gangway.enabled,
            'operators': {op: jobs for op, jobs in sorted(job_map.items())},
            'total_jobs': len(get_all_triggerable_jobs()),
        })

    @app.route('/api/trigger-job', methods=['POST'])
    def api_trigger_job():
        from integrations import get_gangway_client
        gangway = get_gangway_client()
        if not gangway.enabled:
            return jsonify({
                'status': 'disabled',
                'message': 'Gangway not configured. Set PROW_GANGWAY_TOKEN.'
            }), 503

        data = request.get_json(silent=True)
        if not isinstance(data, dict):
            return jsonify({'error': 'Request body must be a JSON object'}), 400
        raw = data.get('job_name') or data.get('operator')
        if not isinstance(raw, str) or not raw.strip():
            return jsonify({'error': 'Missing required field: operator or job_name'}), 400
        job_name_or_operator = raw.strip()

        from integrations.gangway_client import resolve_trigger_target
        resolved_job, operator, resolve_err = resolve_trigger_target(job_name_or_operator)
        if resolve_err:
            return jsonify({'error': resolve_err}), 400

        env_overrides, fbc_err = _parse_fbc_overrides(data)
        if fbc_err:
            return jsonify({'error': fbc_err}), 400

        fbc_display = env_overrides.get('FBC_COMMIT_SHA') or None

        try:
            allowed, remaining, placeholder_id = db.check_cooldown_and_reserve(
                operator, TRIGGER_COOLDOWN_SECONDS,
                fbc_commit_sha=fbc_display, job_name=resolved_job)
        except Exception:
            app.logger.exception("Cooldown check failed for %s", operator)
            return jsonify({'error': 'Rate limit check failed. Try again later.'}), 503
        if not allowed:
            if remaining < 0:
                return jsonify({'error': 'Rate limit check failed. Try again later.'}), 503
            return jsonify({'error': f'Rate limited. Try again in {remaining}s.'}), 429

        result, error = gangway.trigger_job(resolved_job, env_overrides or None)
        if error:
            try:
                db.update_gangway_execution(placeholder_id, "FAILED",
                                            error_message=error)
            except Exception:
                app.logger.exception("DB update failed for %s", resolved_job)
            return jsonify({'error': error}), 502

        execution_id = result.get('execution_id')
        if not execution_id:
            try:
                db.update_gangway_execution(placeholder_id, "FAILED",
                                            error_message="No execution ID returned")
            except Exception:
                app.logger.exception("DB update failed for %s", operator)
            return jsonify({
                'error': 'Gangway returned success but no execution ID',
                'job_name': result.get('job_name'),
            }), 502

        try:
            db.finalize_gangway_execution(
                placeholder_id, execution_id,
                result['job_name'], result['status'])
        except Exception:
            app.logger.exception(
                "Gangway execution %s triggered but DB update failed", execution_id
            )
            return jsonify({
                'error': 'Job was triggered but tracking failed. Do NOT retry.',
                'execution_id': execution_id,
            }), 500

        return jsonify(result)

    @app.route('/api/trigger-all-jobs', methods=['POST'])
    def api_trigger_all_jobs():
        from integrations import get_gangway_client
        from integrations.gangway_client import get_all_triggerable_jobs, operator_from_job_name
        gangway = get_gangway_client()
        if not gangway.enabled:
            return jsonify({
                'status': 'disabled',
                'message': 'Gangway not configured. Set PROW_GANGWAY_TOKEN.'
            }), 503

        data = request.get_json(silent=True) or {}
        env_overrides, fbc_err = _parse_fbc_overrides(data)
        if fbc_err:
            return jsonify({'error': fbc_err}), 400

        fbc_display = env_overrides.get('FBC_COMMIT_SHA') or None

        all_jobs = get_all_triggerable_jobs()
        reserved = []
        skipped = 0

        for job_name in all_jobs:
            operator = operator_from_job_name(job_name) or job_name
            try:
                allowed, remaining, placeholder_id = db.check_cooldown_and_reserve(
                    operator, TRIGGER_COOLDOWN_SECONDS,
                    fbc_commit_sha=fbc_display, job_name=job_name)
            except Exception:
                app.logger.exception("Cooldown check failed for %s", operator)
                continue
            if allowed:
                reserved.append((job_name, operator, placeholder_id))
            else:
                skipped += 1

        if not reserved:
            return jsonify({
                'summary': {'total': len(all_jobs), 'triggered': 0,
                            'skipped': skipped, 'failed': 0},
                'message': f'All {len(all_jobs)} jobs on cooldown',
            })

        def _trigger_in_background(jobs_to_trigger, overrides, gw):
            for idx, (jn, op, pid) in enumerate(jobs_to_trigger):
                if idx > 0:
                    time.sleep(5)
                result, error = gw.trigger_job(jn, overrides or None)
                if error:
                    try:
                        db.update_gangway_execution(pid, "FAILED",
                                                    error_message=error)
                    except Exception:
                        app.logger.exception("DB update failed for %s", jn)
                    continue
                eid = result.get('execution_id')
                if not eid:
                    try:
                        db.update_gangway_execution(pid, "FAILED",
                                                    error_message="No execution ID")
                    except Exception:
                        pass
                    continue
                try:
                    db.finalize_gangway_execution(pid, eid, result['job_name'],
                                                  result['status'])
                except Exception:
                    app.logger.exception("DB finalize failed for %s", eid)

        t = threading.Thread(target=_trigger_in_background,
                             args=(reserved, env_overrides, gangway),
                             daemon=True)
        t.start()

        return jsonify({
            'summary': {
                'total': len(all_jobs),
                'triggered': len(reserved),
                'skipped': skipped,
                'failed': 0,
            },
            'message': f'{len(reserved)} jobs queued for triggering. '
                       f'Click Refresh to see status updates.',
        })

    _TERMINAL_STATUSES = frozenset(('SUCCESS', 'FAILURE', 'ABORTED', 'ERROR'))
    _SKIP_RESOLVE_STATUSES = frozenset(('PENDING', 'QUEUED', 'TRIGGERED', ''))

    def _maybe_resolve_prow_url(prow_url, job_name, triggered_at, status):
        if (not prow_url and job_name
                and (status or '').upper() not in _SKIP_RESOLVE_STATUSES):
            from integrations.gangway_client import GangwayClient
            return GangwayClient.resolve_prow_url(job_name, triggered_at)
        return prow_url

    @app.route('/api/trigger-job/history')
    def api_trigger_job_history():
        operator = request.args.get('operator')
        limit = request.args.get('limit', 20, type=int)
        limit = max(1, min(limit, 100))
        refresh = request.args.get('refresh', '0') == '1'
        if refresh:
            try:
                from integrations.gangway_client import (
                    discover_untracked_builds, get_all_triggerable_jobs,
                    build_spyglass_url,
                )
                known_ids = db.get_tracked_prow_build_ids()
                job_names = get_all_triggerable_jobs()
                discovered = discover_untracked_builds(job_names, known_ids, since_hours=168)
                inserted = 0
                for build in discovered[:20]:
                    prow_url = build_spyglass_url(build['job_name'], build['build_id'])
                    if db.insert_discovered_execution(
                        build['build_id'], build['operator'], build['job_name'],
                        build['result'], build['started'], prow_url,
                    ):
                        inserted += 1
                if inserted:
                    app.logger.info("Discovered %d untracked Prow builds", inserted)
            except Exception:
                app.logger.exception("Failed to discover untracked Prow builds")
            try:
                from integrations.gangway_client import fetch_fbc_sha_from_artifacts
                backfilled = 0
                for ex in db.get_executions_missing_fbc_sha(limit=10):
                    eid = ex['execution_id']
                    bid = eid[5:] if eid.startswith('prow-') else None
                    if not bid:
                        continue
                    sha = fetch_fbc_sha_from_artifacts(ex['job_name'], bid)
                    if sha:
                        db.update_gangway_execution(eid, ex['status'], fbc_commit_sha=sha)
                        backfilled += 1
                if backfilled:
                    app.logger.info("Backfilled FBC SHA for %d builds", backfilled)
            except Exception:
                app.logger.exception("Failed to backfill FBC SHAs")
        executions = db.get_gangway_executions(operator, limit)
        if refresh:
            from integrations import get_gangway_client
            gangway = get_gangway_client()
            if gangway.enabled:
                refreshed = 0
                for ex in executions:
                    if refreshed >= 5:
                        break
                    if (ex.get('status') or '').upper() in _TERMINAL_STATUSES:
                        continue
                    eid = ex.get('execution_id')
                    if not eid:
                        continue
                    refreshed += 1
                    try:
                        remote, err = gangway.get_execution_status(eid)
                        if remote and not err:
                            new_status = remote.get('job_status', ex['status'])
                            prow_url = remote.get('prowjob_url')
                            prow_url = _maybe_resolve_prow_url(
                                prow_url, ex.get('job_name'), ex.get('triggered_at'), new_status)
                            db.update_gangway_execution(eid, new_status, prow_url)
                            ex['status'] = new_status
                            if prow_url:
                                ex['prow_job_url'] = prow_url
                    except Exception:
                        app.logger.exception("Failed to refresh execution %s", eid)
            attempts = 0
            for ex in executions:
                if attempts >= 5:
                    break
                status = (ex.get('status') or '').upper()
                if ex.get('prow_job_url') or not ex.get('job_name') or not ex.get('triggered_at'):
                    continue
                if status not in _TERMINAL_STATUSES:
                    continue
                attempts += 1
                try:
                    url = _maybe_resolve_prow_url(None, ex['job_name'], ex['triggered_at'], status)
                    if url:
                        db.update_gangway_execution(ex['execution_id'], status, url)
                        ex['prow_job_url'] = url
                except Exception:
                    app.logger.exception("Failed to resolve Prow URL for %s", ex.get('execution_id'))
            from integrations.gangway_client import GangwayClient
            reconciled = 0
            for ex in executions:
                if reconciled >= 5:
                    break
                status = (ex.get('status') or '').upper()
                if status in _TERMINAL_STATUSES:
                    continue
                if not ex.get('job_name') or not ex.get('triggered_at'):
                    continue
                reconciled += 1
                try:
                    prow_result, prow_url = GangwayClient.resolve_prow_result(
                        ex['job_name'], ex['triggered_at'])
                    if prow_result:
                        db.update_gangway_execution(ex['execution_id'], prow_result, prow_url)
                        ex['status'] = prow_result
                        if prow_url:
                            ex['prow_job_url'] = prow_url
                except Exception:
                    app.logger.exception("Failed to reconcile trigger %s with Prow",
                                         ex.get('execution_id'))
        resolved_snaps = {}
        for ex in executions:
            sha = ex.get('fbc_commit_sha')
            job_name = ex.get('job_name') or ''
            if sha and _FBC_SHA_RE.fullmatch(sha):
                ver_match = re.search(r'(?:release|main)-(\d+)\.(\d+)', job_name)
                app_name = f"rhwa-fbc-{ver_match.group(1)}{ver_match.group(2)}" if ver_match else None
                cache_key = (sha, app_name or '')
                if cache_key not in resolved_snaps:
                    try:
                        snap_name, _ = _resolve_konflux_snapshot(sha, expected_app=app_name)
                        resolved_snaps[cache_key] = snap_name or ''
                    except Exception:
                        app.logger.warning("Snapshot lookup failed for %s", sha[:8], exc_info=True)
                        resolved_snaps[cache_key] = ''
                ex['snapshot_name'] = resolved_snaps[cache_key]
            else:
                ex['snapshot_name'] = ''
        return jsonify(executions)

    _SAFE_ID = re.compile(r'^[A-Za-z0-9_-]+$')

    @app.route('/api/trigger-job/<execution_id>')
    def api_trigger_job_status(execution_id):
        if not _SAFE_ID.match(execution_id):
            return jsonify({'error': 'Invalid execution id format'}), 400
        from integrations import get_gangway_client
        gangway = get_gangway_client()

        local = db.get_gangway_execution(execution_id)
        if not local:
            return jsonify({'error': 'Execution not found'}), 404

        refresh_error = None
        if gangway.enabled and local['status'] not in _TERMINAL_STATUSES:
            remote, err = gangway.get_execution_status(execution_id)
            if remote and not err:
                new_status = remote.get('job_status', local['status'])
                prow_url = remote.get('prowjob_url')
                prow_url = _maybe_resolve_prow_url(
                    prow_url, local.get('job_name'), local.get('triggered_at'), new_status)
                try:
                    db.update_gangway_execution(execution_id, new_status, prow_url)
                except Exception:
                    app.logger.exception("Failed to update gangway execution %s in DB", execution_id)
                local['status'] = new_status
                if prow_url:
                    local['prow_job_url'] = prow_url
            elif err:
                refresh_error = err

        payload = dict(local)
        if refresh_error:
            payload['refresh_warning'] = f'Could not refresh from Gangway: {refresh_error}'
        return jsonify(payload)

    @app.route('/api/analyze-failure', methods=['POST'])
    def api_analyze_failure():
        """
        Analyze test failure with AI (hybrid: local Claude Code or Anthropic API)
        """
        from ai.analyzer import HybridFailureAnalyzer

        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        test_name = data.get('test_name')
        version = data.get('version')
        platform = data.get('platform')

        if not all([test_name, version, platform]):
            return jsonify({'error': 'Missing required fields: test_name, version, platform'}), 400

        # Check if we already have a recent analysis
        days = data.get('days', 7)
        existing_analysis = db.get_ai_analysis(test_name, version, platform, days)
        if existing_analysis and data.get('use_cached', True):
            existing_analysis['cached'] = True
            return jsonify(existing_analysis)

        # Use provided error_message or get from database
        error_message = data.get('error_message')
        log_url = data.get('log_url', '')

        if not error_message:
            # Get test error details from database
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)

            query = """
                SELECT error_message, log_url
                FROM test_results
                WHERE test_name = ?
                AND version = ?
                AND platform = ?
                AND status = 'failed'
                AND timestamp >= ? AND timestamp <= ?
                ORDER BY timestamp DESC
                LIMIT 1
            """

            cursor = db.conn.cursor()
            cursor.execute(query, (test_name, version, platform,
                                   start_date.isoformat(), end_date.isoformat()))
            test_data = cursor.fetchone()

            if not test_data:
                return jsonify({'error': 'No recent failure found for this test'}), 404

            error_message = test_data[0] or 'No error message'
            log_url = test_data[1] or ''

        # Analyze with hybrid approach
        try:
            analyzer = HybridFailureAnalyzer()
            analysis = analyzer.analyze_failure(
                test_name=test_name,
                error_message=error_message,
                log_url=log_url,
                platform=platform,
                version=version
            )

            # Save analysis to database
            db.save_ai_analysis(test_name, version, platform, analysis)

            analysis['cached'] = False
            return jsonify(analysis)

        except Exception as e:
            return jsonify({
                'error': f'Analysis failed: {str(e)}',
                'root_cause': 'Analysis service error',
                'confidence': 0
            }), 500

    @app.route('/api/analysis-stats')
    def api_analysis_stats():
        """Get statistics about AI analyses"""
        stats = db.get_analysis_stats()
        return jsonify(stats)

    @app.route('/api/save-classification', methods=['POST'])
    def api_save_classification():
        """
        Save manual classification for a test failure
        """
        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        test_name = data.get('test_name')
        version = data.get('version')
        platform = data.get('platform')
        classification = data.get('classification')

        if not all([test_name, version, platform, classification]):
            return jsonify({'error': 'Missing required fields: test_name, version, platform, classification'}), 400

        # Validate classification
        valid_classifications = ['product_bug', 'automation_bug', 'system_issue', 'transient', 'to_investigate']
        if classification not in valid_classifications:
            return jsonify({'error': f'Invalid classification. Must be one of: {", ".join(valid_classifications)}'}), 400

        # Save to database
        rows_updated = db.save_manual_classification(
            test_name=test_name,
            version=version,
            platform=platform,
            classification=classification,
            classified_by='user'
        )

        if rows_updated > 0:
            return jsonify({
                'status': 'success',
                'rows_updated': rows_updated,
                'classification': classification
            })
        else:
            return jsonify({'error': 'No matching test result found to update'}), 404

    @app.route('/api/get-test-data', methods=['POST'])
    def api_get_test_data():
        """
        Get existing data for a test (classification, Jira key, AI analysis)
        """
        data = request.json
        if not data:
            return jsonify({'error': 'Missing request data'}), 400

        test_name = data.get('test_name')
        version = data.get('version')
        platform = data.get('platform')

        if not all([test_name, version, platform]):
            return jsonify({'error': 'Missing required fields: test_name, version, platform'}), 400

        result = {
            'manual_classification': None,
            'jira_issue_key': None,
            'ai_analysis': None
        }

        # Get manual classification and Jira issue from test_results
        cursor = db.conn.cursor()

        # Log query parameters for debugging
        logger.info(f"Fetching test data: test_name={test_name}, version={version}, platform={platform}")

        cursor.execute("""
            SELECT manual_classification, jira_issue_key
            FROM test_results
            WHERE test_name = ?
            AND version = ?
            AND UPPER(platform) = UPPER(?)
            AND status = 'failed'
            ORDER BY timestamp DESC
            LIMIT 1
        """, (test_name, version, platform))

        row = cursor.fetchone()
        if row:
            result['manual_classification'] = row[0]
            result['jira_issue_key'] = row[1]
            logger.info(f"Found test data: classification={row[0]}, jira_key={row[1]}")
        else:
            logger.info(f"No test data found for {test_name}/{version}/{platform}")

        # Get AI analysis
        ai_analysis = db.get_ai_analysis(test_name, version, platform, days=90)
        if ai_analysis:
            result['ai_analysis'] = ai_analysis

        return jsonify(result)

    @app.route('/report')
    def report_page():
        """Shareable, print-friendly report page. Filters via query params."""
        days = request.args.get('days', 7, type=int)
        operator = request.args.get('operator')
        version = normalize_version(request.args.get('version'))

        summary = calculator.get_summary_stats(days=days, version=version)
        op_stats = db.get_operator_stats(days=days, version=version)
        top_tests = calculator.get_test_rankings(days=days, version=version, limit=20)
        regressions_data = db.get_regressions(operator=operator, version=version)
        job_runs = db.get_job_run_history(days=days, operator=operator, version=version)

        if operator:
            top_tests = [t for t in top_tests if (t.get('test_name', '').upper().startswith(operator.upper()) or
                          operator.upper() in (t.get('test_name', '') or '').upper())]
            op_stats = [s for s in op_stats if s.get('operator', '').upper() == operator.upper()]

        reg_list = [r for r in regressions_data if r.get('change_type') == 'regression']
        fix_list = [r for r in regressions_data if r.get('change_type') == 'fix']
        failing = [t for t in top_tests if t.get('pass_rate', 100) < 100]

        now = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
        start = (datetime.utcnow() - timedelta(days=days)).strftime('%Y-%m-%d')
        end = datetime.utcnow().strftime('%Y-%m-%d')
        title = (operator or 'medik8s') + ' CI Report'
        subtitle = start + ' to ' + end + ' (' + str(days) + ' days)'
        if version:
            subtitle += ' | OCP ' + version

        def _clean(name):
            if not name:
                return ''
            for pfx in ['[It] ', 'FAR ', 'SBR ', 'SNR ', 'NHC ', 'MDR ', 'NMO ']:
                if name.startswith(pfx):
                    name = name[len(pfx):]
            return name

        pass_rate = summary.get('avg_pass_rate', 0)
        rate_cls = 'green' if pass_rate >= 90 else 'red'

        ops_html = ''
        if len(op_stats) > 1:
            rows = ''.join(
                '<tr><td><strong>' + str(s.get('operator', '?')) + '</strong></td>'
                '<td>' + str(s.get('total_tests', 0)) + '</td>'
                '<td>' + str(s.get('passed', 0)) + '</td>'
                '<td>' + str(s.get('failed', 0)) + '</td>'
                '<td>' + str(s.get('pass_rate', 0)) + '%</td></tr>'
                for s in op_stats
            )
            ops_html = '<h2>Operator Breakdown</h2><table><tr><th>Operator</th><th>Tests</th><th>Passed</th><th>Failed</th><th>Pass Rate</th></tr>' + rows + '</table>'

        def _change_table(items, heading, badge_cls, label):
            if not items:
                return ''
            rows = ''.join(
                '<tr><td>' + _clean(r.get('test_name', '')) + '</td>'
                '<td>' + str(r.get('operator', '')) + '</td>'
                '<td><span class="badge ' + badge_cls + '">' + label + '</span></td></tr>'
                for r in items
            )
            return '<h2>' + heading + ' (' + str(len(items)) + ')</h2><table><tr><th>Test</th><th>Operator</th><th>Status</th></tr>' + rows + '</table>'

        reg_html = _change_table(reg_list, 'Regressions', 'badge-fail', 'passed &rarr; failed')
        fix_html = _change_table(fix_list, 'Fixes', 'badge-pass', 'failed &rarr; passed')

        fail_html = ''
        if failing:
            rows = ''.join(
                '<tr><td>' + _clean(t.get('test_name', '')) + '</td>'
                '<td><strong class="red">' + '{:.1f}'.format(t.get('pass_rate', 0)) + '%</strong></td>'
                '<td>' + str(t.get('total_runs', 0)) + '</td></tr>'
                for t in failing[:10]
            )
            fail_html = '<h2>Failing Tests</h2><table><tr><th>Test</th><th>Pass Rate</th><th>Runs</th></tr>' + rows + '</table>'
        else:
            fail_html = '<p style="color:#73bf69;">All tests passing at 100%.</p>'

        job_rows = ''
        for r in job_runs[:15]:
            jn = r.get('job_name', '') or ''
            short = jn.split('e2e-')[-1] if 'e2e-' in jn else jn[:40]
            st = r.get('status', '?')
            bcls = 'badge-pass' if st == 'passed' else 'badge-fail'
            rd = (r.get('run_date', '') or '')[:10]
            job_rows += ('<tr><td>' + short + '</td><td><span class="badge ' + bcls + '">'
                         + st + '</span></td><td>' + rd + '</td><td>'
                         + str(r.get('passed_tests', 0)) + '/' + str(r.get('total_tests', 0))
                         + '</td><td>' + str(r.get('ocp_version', '') or '') + '</td></tr>')

        permalink = '/report?days=' + str(days)
        if operator:
            permalink += '&operator=' + operator
        if version:
            permalink += '&version=' + version

        css = """body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; max-width: 900px; margin: 0 auto; padding: 20px; color: #d8d9da; background: #111217; }
h1 { font-size: 22px; margin: 0; color: #f4f5f5; } h2 { font-size: 16px; margin: 20px 0 10px; border-bottom: 2px solid #2c3235; padding-bottom: 6px; color: #f4f5f5; }
.subtitle { color: #9fa7b3; font-size: 14px; margin: 4px 0 20px; }
.stats { display: flex; gap: 16px; margin-bottom: 20px; flex-wrap: wrap; }
.stat { background: #1a1b21; border-radius: 8px; padding: 14px 20px; text-align: center; min-width: 120px; border: 1px solid #2c3235; }
.stat .value { font-size: 28px; font-weight: 700; } .stat .label { font-size: 11px; color: #9fa7b3; text-transform: uppercase; }
.green { color: #73bf69; } .red { color: #f2495c; } .blue { color: #5794f2; }
table { width: 100%; border-collapse: collapse; font-size: 13px; margin-bottom: 20px; }
th { background: #22252b; text-align: left; padding: 8px 10px; border: 1px solid #2c3235; font-weight: 600; color: #d8d9da; }
td { padding: 8px 10px; border: 1px solid #2c3235; }
.badge { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }
.badge-pass { background: rgba(115,191,105,0.15); color: #73bf69; } .badge-fail { background: rgba(242,73,92,0.15); color: #f2495c; }
.footer { margin-top: 30px; padding-top: 10px; border-top: 1px solid #2c3235; font-size: 11px; color: #9fa7b3; }
a { color: #5794f2; text-decoration: none; } a:hover { text-decoration: underline; }
@media print { body { padding: 0; background: #111217; } .no-print { display: none; } }"""

        html = ('<!DOCTYPE html><html><head><meta charset="UTF-8"><title>' + title
                + '</title><style>' + css + '</style></head><body>'
                + '<h1>' + title + '</h1>'
                + '<p class="subtitle">' + subtitle + ' | Generated: ' + now + '</p>'
                + '<div class="stats">'
                + '<div class="stat"><div class="value">' + str(summary.get('total_tests', 0)) + '</div><div class="label">Tests</div></div>'
                + '<div class="stat"><div class="value ' + rate_cls + '">' + '{:.1f}'.format(pass_rate) + '%</div><div class="label">Pass Rate</div></div>'
                + '<div class="stat"><div class="value red">' + str(summary.get('failed_tests', 0)) + '</div><div class="label">Failures</div></div>'
                + '<div class="stat"><div class="value blue">' + str(len(job_runs)) + '</div><div class="label">Job Runs</div></div>'
                + '<div class="stat"><div class="value red">' + str(len(reg_list)) + '</div><div class="label">Regressions</div></div>'
                + '<div class="stat"><div class="value green">' + str(len(fix_list)) + '</div><div class="label">Fixes</div></div>'
                + '</div>'
                + ops_html + reg_html + fix_html + fail_html
                + '<h2>Recent Job Runs</h2>'
                + '<table><tr><th>Job</th><th>Status</th><th>Date</th><th>Tests</th><th>OCP</th></tr>'
                + job_rows + '</table>'
                + '<div class="footer"><a href="/" class="no-print">Back to Dashboard</a> | '
                + 'medik8s CI Dashboard | <a href="' + permalink + '">Permalink</a></div>'
                + '</body></html>')
        return html, 200, {'Content-Type': 'text/html'}

    @app.route('/api/export')
    def api_export():
        """Export test results to XLSX, CSV, or MD format matching reference Google Sheet"""
        export_format = request.args.get('format', 'xlsx')
        days = request.args.get('days', 30, type=int)
        version_param = request.args.get('version')
        version = normalize_version(version_param)

        logger.info(f"[EXPORT] Received: format={export_format}, days={days}, version_param={version_param}, normalized_version={version}")

        enriched_rows = db.get_enriched_test_results(days=days, version=version)
        logger.info(f"[EXPORT] Found {len(enriched_rows)} enriched test results")

        today = datetime.now().strftime('%Y-%m-%d')
        filename = f'dashboard-export-{version}-{days}days-{today}'

        if export_format != 'xlsx':
            return jsonify({'error': 'Invalid format. Only xlsx is supported'}), 400
        return export_to_xlsx_enriched(enriched_rows, filename, version, days)

    def export_to_xlsx_enriched(enriched_rows, filename, version, days):
        """Export to Excel matching the reference Google Sheet 17-column structure"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Test Results'

        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        link_font = Font(color='0563C1', underline='single', size=10)
        data_font = Font(size=10)
        pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        pass_font = Font(color='006100', size=10, bold=True)
        fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        fail_font = Font(color='9C0006', size=10, bold=True)
        alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

        headers = [
            'Test Name', 'Operator', 'Result', 'Periodic Job', 'Run Date',
            'Job Duration', 'OCP Version', 'Platform', 'Operator CSV Version',
            'FBC Catalog Image', 'Prow Job', 'E2E Test Log',
            'Operator Install Log', 'CatalogSource Log', 'Artifacts', 'Build Log', 'Polarion ID'
        ]
        col_widths = [48, 11, 10, 30, 14, 14, 44, 11, 37, 42, 13, 16, 24, 24, 15, 13, 15]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.row_dimensions[1].height = 35
        ws.freeze_panes = 'A2'

        for col_num, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        for row_idx, row in enumerate(enriched_rows, 2):
            fmt = _format_export_row(row)
            short_job = fmt['short_job']
            duration_str = fmt['duration_str']
            run_date = fmt['run_date']
            result_str = fmt['result_str']

            fbc_image = row.get('fbc_image') or ''
            fbc_short = _fbc_short(fbc_image)

            polarion_id = row.get('polarion_id') or ''

            is_alt = (row_idx % 2 == 0)

            ws.cell(row=row_idx, column=1, value=row.get('test_description') or row.get('test_name')).font = data_font
            ws.cell(row=row_idx, column=2, value=row.get('operator') or '').font = data_font
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')

            result_cell = ws.cell(row=row_idx, column=3, value=result_str)
            result_cell.alignment = Alignment(horizontal='center')
            if result_str == 'PASSED':
                result_cell.fill = pass_fill
                result_cell.font = pass_font
            elif result_str == 'FAILED':
                result_cell.fill = fail_fill
                result_cell.font = fail_font

            ws.cell(row=row_idx, column=4, value=short_job).font = data_font
            ws.cell(row=row_idx, column=5, value=run_date).font = data_font
            ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal='center')
            ws.cell(row=row_idx, column=6, value=duration_str).font = data_font
            ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal='center')
            ws.cell(row=row_idx, column=7, value=row.get('ocp_version') or row.get('version') or '').font = data_font
            ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal='center')
            ws.cell(row=row_idx, column=8, value=(row.get('platform') or '').upper()).font = data_font
            ws.cell(row=row_idx, column=8).alignment = Alignment(horizontal='center')
            ws.cell(row=row_idx, column=9, value=row.get('csv_version') or '').font = data_font

            fbc_cell = ws.cell(row=row_idx, column=10, value=fbc_short or '-')
            if fbc_image and 'quay.io/' in fbc_image:
                repo_path = fbc_image.split('quay.io/')[-1].split('@')[0].split(':')[0]
                fbc_url = f"https://quay.io/repository/{repo_path}?tab=tags"
                fbc_cell.hyperlink = fbc_url
                fbc_cell.font = link_font

            prow_url = row.get('job_url') or ''
            prow_cell = ws.cell(row=row_idx, column=11, value='View Job' if prow_url else '-')
            prow_cell.alignment = Alignment(horizontal='center')
            if prow_url:
                prow_cell.hyperlink = prow_url
                prow_cell.font = link_font

            e2e_url = fmt['e2e_log_url']
            e2e_cell = ws.cell(row=row_idx, column=12, value='Test Log' if e2e_url else '-')
            e2e_cell.alignment = Alignment(horizontal='center')
            if e2e_url:
                e2e_cell.hyperlink = e2e_url
                e2e_cell.font = link_font

            sub_url = fmt['subscribe_log_url']
            sub_cell = ws.cell(row=row_idx, column=13, value='Install Log' if sub_url else '-')
            sub_cell.alignment = Alignment(horizontal='center')
            if sub_url:
                sub_cell.hyperlink = sub_url
                sub_cell.font = link_font

            cat_url = fmt['catalog_log_url']
            cat_cell = ws.cell(row=row_idx, column=14, value='Catalog Log' if cat_url else '-')
            cat_cell.alignment = Alignment(horizontal='center')
            if cat_url:
                cat_cell.hyperlink = cat_url
                cat_cell.font = link_font

            art_url = fmt['artifacts_url']
            art_cell = ws.cell(row=row_idx, column=15, value='Artifacts' if art_url else '-')
            art_cell.alignment = Alignment(horizontal='center')
            if art_url:
                art_cell.hyperlink = art_url
                art_cell.font = link_font

            bld_url = fmt['build_log_url']
            bld_cell = ws.cell(row=row_idx, column=16, value='Build Log' if bld_url else '-')
            bld_cell.alignment = Alignment(horizontal='center')
            if bld_url:
                bld_cell.hyperlink = bld_url
                bld_cell.font = link_font

            pol_cell = ws.cell(row=row_idx, column=17, value=polarion_id or '-')
            pol_cell.alignment = Alignment(horizontal='center')
            if polarion_id:
                pol_cell.hyperlink = _polarion_url(polarion_id)
                pol_cell.font = link_font

            if is_alt:
                for c in range(1, 18):
                    if c == 3:  # skip Result column to preserve pass/fail coloring
                        continue
                    ws.cell(row=row_idx, column=c).fill = alt_fill

        ws.auto_filter.ref = f"A1:Q{len(enriched_rows) + 1}"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{filename}.xlsx'
        )

    @app.teardown_appcontext
    def close_db(error):
        """Close database connection on app shutdown"""
        if error:
            print(f"App error: {error}")

    return app
